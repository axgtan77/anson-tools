"""Convert EV Charging.xlsx into seed.json for the charging-log web app."""
import json
from datetime import datetime, date
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "EV Charging.xlsx"
OUT = ROOT / "charging-log" / "seed.json"

HEADERS = [
    "date", "vehicle", "odo",
    "start_pct", "target_pct", "needed_pct", "kwh_needed",
    "charger", "est_hours", "time_start", "est_end", "actual_end",
    "actual_final_pct", "actual_charged_pct", "actual_kwh",
    "actual_hours", "actual_kw", "location", "cost_php",
    "km_per_kwh", "km_since_last", "days_since_last", "notes",
]


def iso(v):
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return v


def clean(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return iso(v)
    if isinstance(v, float):
        return round(v, 4)
    return v


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    charging = []
    ws = wb["EV Charging"]
    rows = list(ws.iter_rows(values_only=True))
    for r in rows[1:]:
        if not r or r[0] is None:
            continue
        row = list(r[:23])
        while len(row) < 23:
            row.append(None)
        entry = dict(zip(HEADERS, [clean(v) for v in row]))
        if entry["date"]:
            entry["id"] = f"seed-{len(charging):03d}"
            charging.append(entry)

    vehicles = []
    ws = wb["EVs"]
    for r in list(ws.iter_rows(values_only=True))[1:]:
        if not r or not r[0]:
            continue
        vehicles.append({
            "model": r[0],
            "battery_kwh": clean(r[1]),
            "range_km": clean(r[3]),
        })

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w") as f:
        json.dump({"vehicles": vehicles, "charging": charging}, f, indent=2)

    print(f"Wrote {len(charging)} charges + {len(vehicles)} vehicles → {OUT}")


if __name__ == "__main__":
    main()
